from typing import Any
import ass, re 
from datetime import timedelta
import openpyxl
import xlrd

DEBUG = False

def __fmt(s) -> str:
    if s is None:
        return ""    
    elif type(s) == str:
        return s.strip()
    elif type(s) == int:
        return str(s)
    elif type(s) == float:
        return str(s)
    else:
        return str(s)

def __load_file_openpyxl(fn: str) -> dict[str, list[list[str]]]:
    book = openpyxl.load_workbook(fn)
    return dict({ws: list([list([__fmt(c) for c in row]) for row in book[ws].iter_rows(values_only=True)]) for ws in book.sheetnames})

def __load_file_xlrd(fn: str) -> dict[str, list[list[str]]]:
    book = xlrd.open_workbook(fn)
    return dict({ws_name: list([list([__fmt(c.value) for c in row]) for row in book.sheet_by_name(ws_name).get_rows()]) for ws_name in book.sheet_names()})

def load_excel_file(fn: str) -> dict[str, list[list[str]]]:
    if DEBUG:
        print("Loading file...")
    try:
        data = __load_file_openpyxl(fn) 
        if DEBUG:
            print("OpenPYXL loaded the file successfully!")
        return data
    except:
        if DEBUG:
            print("OpenPYXL refused. Trying xlrd...")
    
    try:
        data = __load_file_xlrd(fn)
        if DEBUG:
            print("xlrd loaded the file successfully!")
        return data
    except:
        raise Exception(f"No valid interpreter found for {fn}")

DATETIME_REGEX = re.compile(r"(\d+):(\d\d):(\d\d)[:.](\d+)")
def is_timestamp(s: str) -> bool:
    return re.match(DATETIME_REGEX, s) is not None

def convert_datetime(s: str, is_timecode: bool = True, framerate: float = 24.0, shift: timedelta = None, scale: float = None):
    is_negative = False
    if s[0] == "-":
        is_negative = True
        s = s[1:]

    m = re.match(DATETIME_REGEX, s)

    if not m:
        raise ValueError(f"Invalid timestamp format: {s}")

    hours = int(m[1])
    minutes = int(m[2])
    seconds = int(m[3])
    if is_timecode:
        centiseconds = int(int(m[4])/framerate * 100)
    else:
        centiseconds = int(int(m[4])/pow(10.0, len(m[3])-2))
    td = timedelta(hours=hours, minutes=minutes, seconds=(seconds+centiseconds/100.0))
    if shift:
        td = td + shift
    if scale:
        td = td * scale

    return td

def create_style(name: str = "Default", is_type: bool = False) -> ass.Style:
    default_style = ass.Style()
    default_style.name = name
    default_style.fontname = "Trebuchet MS"
    default_style.fontsize = "24"
    default_style.shadow = "2"
    default_style.margin_l = "40"
    default_style.margin_r = "40"
    default_style.margin_v = "20"
    if is_type:
        default_style.alignment = "8"
    return default_style

def create_document() -> ass.Document:
    doc = ass.Document()
    doc.script_type = "v4.00+"
    doc.sections["Script Info"]["WrapStyle"] = "0"
    doc.sections["Script Info"]["ScaledBorderAndShadow"] = "yes"
    doc.sections["Script Info"]["YCbCr Matrix"] = "TV.601"
    doc.sections["Script Info"]["PlayResX"] = "640"
    doc.sections["Script Info"]["PlayResY"] = "360"
    doc.styles.append(create_style())
    return doc

def find_style(doc: ass.Document, style_name: str) -> ass.Style | None:
    for s in doc.styles:
        if s.name == style_name:
            return s
    return None

def convert_worksheet_to_ass(ws: list[list[str]], doc: ass.Document = None, start_col: int = None, end_col: int = None, dialogue_col: int = None, 
                            actor_col: int = None, track_col: int = None, italics_col: int = None, has_headers: bool = True,
                            convert_timestamp_args: dict[str, Any] = {}) -> ass.Document:
    if start_col is None and dialogue_col is None:
        raise ValueError("Inform at least one of the following: start_col, dialogue_col")
    
    if not doc:
        doc = create_document()
    
    skip_headers = has_headers
    for row in ws:
        if skip_headers:
            skip_headers = False
            continue
        event = ass.Dialogue()
        if start_col:
            event.start = convert_datetime(row[start_col], **convert_timestamp_args)
        else:
            event.start = "0:00:00.00"

        if end_col:
            event.end = convert_datetime(row[end_col], **convert_timestamp_args)
        elif start_col:
            event.end = event.start
        else:
            event.end = "0:00:00.00"
        
        if dialogue_col:
            event.text = row[dialogue_col].replace("\n", "\\N")

        if italics_col:
            if row[italics_col]:
                event.text = "{\\i1}" + event.text

        if actor_col:
            event.name = row[actor_col]

        if track_col:
            track = row[track_col]
            if not find_style(doc, track):
                doc.styles.append(create_style(track))
            event.style = track
        else:
            event.style = "Default"
        doc.events.append(event)

    return doc

if __name__ == "__main__":
    input_fn = "test.xlsx"
    output_fn = "test_xlsx.ass"

    wb = load_excel_file(input_fn)    
    ws = wb[list(wb.keys())[0]]
    doc = convert_worksheet_to_ass(ws, start_col=0, end_col=1, actor_col=2, dialogue_col=3, has_headers=True)
    with open(output_fn, "w", encoding="utf_8_sig") as f:
        doc.dump_file(f)

    input_fn = "test.xls"
    output_fn = "test_xls.ass"

    wb = load_excel_file(input_fn)    
    ws = wb[list(wb.keys())[0]]
    doc = convert_worksheet_to_ass(ws, track_col=1, start_col=2, end_col=3, dialogue_col=6, has_headers=True)
    with open(output_fn, "w", encoding="utf_8_sig") as f:
        doc.dump_file(f)